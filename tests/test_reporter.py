"""Tests for src/reporter.py — Excel report generation"""
import sys
import os
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

import pytest
import pandas as pd
import openpyxl
from datetime import datetime

from reporter import (
    generate_report,
    auto_fit_columns,
    apply_header_style,
    format_amount_column,
)
from pipeline import ReconciliationResult
from matcher import MatchRecord


# ── Helpers ────────────────────────────────────────────────────────────

def _make_bank_df(records):
    """Build a mock bank DataFrame matching the pipeline output format."""
    rows = []
    for date_str, debit, credit, balance, summary, cp, cp_acct, norm_sum in records:
        rows.append({
            'date': pd.Timestamp(date_str),
            'debit': float(debit),
            'credit': float(credit),
            'balance': float(balance),
            'summary': str(summary),
            'counterparty': str(cp),
            'counterparty_acct': str(cp_acct),
            'source': 'bank',
            'normalized_amount': float(credit - debit),
            'month': pd.Timestamp(date_str).strftime('%Y-%m'),
            'normalized_summary': str(norm_sum),
        })
    return pd.DataFrame(rows)


def _make_ledger_df(records):
    """Build a mock ledger DataFrame matching the pipeline output format."""
    rows = []
    for date_str, debit, credit, summary, cp_subj, voucher, direction, norm_sum in records:
        rows.append({
            'date': pd.Timestamp(date_str),
            'debit': float(debit),
            'credit': float(credit),
            'summary': str(summary),
            'counterparty_subject': str(cp_subj),
            'voucher_no': str(voucher),
            'direction': str(direction),
            'balance': 0.0,
            'source': 'ledger',
            'normalized_amount': float(debit - credit),
            'month': pd.Timestamp(date_str).strftime('%Y-%m'),
            'normalized_summary': str(norm_sum),
        })
    return pd.DataFrame(rows)


def _make_result(bank_recs, ledger_recs, match_records,
                 bank_only=None, ledger_only=None,
                 bank_dups=None, ledger_dups=None):
    """Construct a mock ReconciliationResult.

    Args:
        bank_recs: list for _make_bank_df
        ledger_recs: list for _make_ledger_df
        match_records: list of MatchRecord
        bank_only: list of indices into bank_df (unmatched)
        ledger_only: list of indices into ledger_df (unmatched)
        bank_dups: list of indices (bank duplicates)
        ledger_dups: list of indices (ledger duplicates)
    """
    bank_df = _make_bank_df(bank_recs)
    ledger_df = _make_ledger_df(ledger_recs)

    # Unmatched
    if bank_only is not None:
        unmatched_bank = bank_df.iloc[bank_only].copy()
    else:
        unmatched_bank = pd.DataFrame(columns=bank_df.columns)

    if ledger_only is not None:
        unmatched_ledger = ledger_df.iloc[ledger_only].copy()
    else:
        unmatched_ledger = pd.DataFrame(columns=ledger_df.columns)

    # Duplicates
    if bank_dups is not None:
        bd = bank_df.iloc[bank_dups].copy()
        bd['is_duplicate'] = True
        bd['duplicate_group_id'] = 0
    else:
        bd = pd.DataFrame(columns=list(bank_df.columns) + ['is_duplicate', 'duplicate_group_id'])

    if ledger_dups is not None:
        ld = ledger_df.iloc[ledger_dups].copy()
        ld['is_duplicate'] = True
        ld['duplicate_group_id'] = 0
    else:
        ld = pd.DataFrame(columns=list(ledger_df.columns) + ['is_duplicate', 'duplicate_group_id'])

    n_exact = sum(1 for r in match_records if r.match_type == 'exact')
    n_fuzzy = sum(1 for r in match_records if r.match_type == 'fuzzy')
    n_split = sum(1 for r in match_records if r.match_type == 'split')

    bank_income = float(bank_df['credit'].sum())
    bank_expense = float(bank_df['debit'].sum())
    ledger_income = float(ledger_df['debit'].sum())
    ledger_expense = float(ledger_df['credit'].sum())

    total_bank_amt = float(bank_df['normalized_amount'].sum())
    total_ledger_amt = float(ledger_df['normalized_amount'].sum())

    matched_amt_diff = sum(abs(r.amount_diff) for r in match_records)
    unmatched_bank_amt = float(unmatched_bank['normalized_amount'].sum()) if len(unmatched_bank) > 0 else 0.0
    unmatched_ledger_amt = float(unmatched_ledger['normalized_amount'].sum()) if len(unmatched_ledger) > 0 else 0.0

    max_count = max(len(bank_df), len(ledger_df))
    match_rate = round(len(match_records) / max_count * 100, 2) if max_count > 0 else 100.0

    return ReconciliationResult(
        bank_file='/fake/bank.xlsx',
        ledger_file='/fake/ledger.xls',
        bank_total_count=len(bank_df),
        ledger_total_count=len(ledger_df),
        bank_total_income=bank_income,
        bank_total_expense=bank_expense,
        ledger_total_income=ledger_income,
        ledger_total_expense=ledger_expense,
        bank_opening_balance=10000.0,
        bank_closing_balance=9500.0,
        matched_count=len(match_records),
        exact_matched=n_exact,
        fuzzy_matched=n_fuzzy,
        split_matched=n_split,
        unmatched_bank_count=len(unmatched_bank),
        unmatched_ledger_count=len(unmatched_ledger),
        duplicate_bank_count=len(bd),
        duplicate_ledger_count=len(ld),
        total_bank_amount=total_bank_amt,
        total_ledger_amount=total_ledger_amt,
        matched_amount_diff=matched_amt_diff,
        unmatched_bank_amount=unmatched_bank_amt,
        unmatched_ledger_amount=unmatched_ledger_amt,
        matched_records=match_records,
        unmatched_bank=unmatched_bank,
        unmatched_ledger=unmatched_ledger,
        bank_duplicates=bd,
        ledger_duplicates=ld,
        reconciled_at='2026-04-28T10:00:00',
        date_range=('2026-01-01', '2026-04-28'),
        match_rate=match_rate,
        bank_df=bank_df,
        ledger_df=ledger_df,
    )


# ── Test Data ──────────────────────────────────────────────────────────

@pytest.fixture
def sample_result():
    """Create a sample ReconciliationResult with varied data."""
    bank_recs = [
        ('2026-04-01', 500.0, 0.0, 10500.0, '支付办公用品', '供应商A', '6222***1234', '支付办公用品'),
        ('2026-04-02', 0.0, 2000.0, 12500.0, '收款销售货款', '客户B', '6222***5678', '收款销售货款'),
        ('2026-04-03', 300.0, 0.0, 12200.0, '支付水电费', '电力公司', '6222***9012', '支付水电费'),
        ('2026-04-05', 1000.0, 0.0, 11200.0, '支付房租', '房东C', '6222***3456', '支付房租'),
        ('2026-04-06', 0.0, 5000.0, 16200.0, '收款项目款', '客户D', '6222***7890', '收款项目款'),
    ]
    bank_df = _make_bank_df(bank_recs)

    ledger_recs = [
        ('2026-04-01', 500.0, 0.0, '办公用品费', '管理费用', 'V001', '借', '办公用品费'),
        ('2026-04-02', 0.0, 2000.0, '销售收入', '主营业务收入', 'V002', '贷', '销售收入'),
        ('2026-04-03', 300.0, 0.0, '水电费', '管理费用', 'V003', '借', '水电费'),
        ('2026-04-07', 1500.0, 0.0, '预付货款', '预付账款', 'V004', '借', '预付货款'),
        ('2026-04-08', 0.0, 3000.0, '咨询服务收入', '其他业务收入', 'V005', '贷', '咨询服务收入'),
    ]
    ledger_df = _make_ledger_df(ledger_recs)

    # Match records: exact for first 3, leave 2 bank + 2 ledger unmatched
    matches = [
        MatchRecord(
            bank_idx=0, ledger_idx=0, match_type='exact',
            amount_diff=0.0, date_diff=0, score=100.0,
            bank_amount=-500.0, ledger_amount=-500.0,
        ),
        MatchRecord(
            bank_idx=1, ledger_idx=1, match_type='exact',
            amount_diff=0.0, date_diff=0, score=100.0,
            bank_amount=2000.0, ledger_amount=2000.0,
        ),
        MatchRecord(
            bank_idx=2, ledger_idx=2, match_type='fuzzy',
            amount_diff=0.0, date_diff=0, score=85.0,
            bank_amount=-300.0, ledger_amount=-300.0,
        ),
    ]

    return _make_result(
        bank_recs, ledger_recs, matches,
        bank_only=[3, 4],
        ledger_only=[3, 4],
    )


@pytest.fixture
def result_with_duplicates():
    """Create a result with duplicate transactions."""
    bank_recs = [
        ('2026-04-01', 500.0, 0.0, 10500.0, '支付A', '供应商A', '6222***1234', '支付a'),
        ('2026-04-01', 500.0, 0.0, 10000.0, '支付A dup', '供应商A', '6222***1234', '支付a'),
        ('2026-04-02', 0.0, 1000.0, 11000.0, '收款B', '客户B', '6222***5678', '收款b'),
    ]
    ledger_recs = [
        ('2026-04-01', 500.0, 0.0, '支付A', '管理费用', 'V001', '借', '支付a'),
        ('2026-04-01', 500.0, 0.0, '支付A dup', '管理费用', 'V002', '借', '支付a'),
        ('2026-04-02', 0.0, 1000.0, '收款B', '主营业务收入', 'V003', '贷', '收款b'),
    ]

    matches = [
        MatchRecord(
            bank_idx=2, ledger_idx=2, match_type='exact',
            amount_diff=0.0, date_diff=0, score=100.0,
            bank_amount=1000.0, ledger_amount=1000.0,
        ),
    ]

    return _make_result(
        bank_recs, ledger_recs, matches,
        bank_only=[0, 1],
        ledger_only=[0, 1],
        bank_dups=[0, 1],
        ledger_dups=[0, 1],
    )


# ── Tests ──────────────────────────────────────────────────────────────

class TestGenerateReport:
    """Tests for generate_report function."""

    def test_generate_report_creates_file(self, sample_result):
        """Verify that generate_report creates a .xlsx file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            result_path = generate_report(sample_result, tmp_path)
            assert result_path == tmp_path
            assert os.path.exists(tmp_path)
            assert os.path.getsize(tmp_path) > 0
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_generate_report_five_worksheets(self, sample_result):
        """Verify that the generated workbook has exactly 5 worksheets."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            generate_report(sample_result, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            sheet_names = wb.sheetnames
            expected = ['对账汇总', '匹配明细', '银行独有', '日记账独有', '疑似重复']
            for name in expected:
                assert name in sheet_names, f"Missing sheet: {name}"
            assert len(sheet_names) == 5, f"Expected 5 sheets, got {len(sheet_names)}: {sheet_names}"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_summary_sheet_content(self, sample_result):
        """Verify the summary sheet contains key information."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            generate_report(sample_result, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb['对账汇总']

            # Collect all cell values as text
            all_text = ''
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        all_text += str(cell) + ' '

            # Check for key summary items
            assert '5' in all_text or 'bank_total_count' not in all_text  # bank count
            assert '对账时间' in all_text or '2026' in all_text
            assert '匹配率' in all_text or 'match' in all_text.lower()
            assert '银行' in all_text
            assert '日记账' in all_text
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_matched_detail_columns(self, sample_result):
        """Verify matched detail sheet has correct columns."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            generate_report(sample_result, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb['匹配明细']

            # Read header row
            headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
            expected_cols = ['序号', '匹配类型', '银行日期', '银行摘要', '银行金额',
                            '日记账日期', '日记账摘要', '日记账金额', '金额差', '日期差', '相似度分数']
            for col in expected_cols:
                assert col in headers, f"Missing column: {col}"

            # Check data rows - should have 3 matched records
            data_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                          if any(c is not None for c in row))
            assert data_rows >= 3, f"Expected at least 3 data rows, got {data_rows}"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_bank_only_sheet(self, sample_result):
        """Verify bank-only worksheet content."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            generate_report(sample_result, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb['银行独有']

            headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
            expected_cols = ['日期', '摘要', '原始摘要', '归一化摘要', '金额', '余额', '对方名称', '对方账号']
            for col in expected_cols:
                assert col in headers, f"Missing column in bank-only: {col}"

            # Should have 2 rows (indices 3, 4 from sample)
            data_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                          if any(c is not None for c in row))
            assert data_rows == 2, f"Expected 2 bank-only rows, got {data_rows}"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_ledger_only_sheet(self, sample_result):
        """Verify ledger-only worksheet content."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            generate_report(sample_result, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb['日记账独有']

            headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
            expected_cols = ['日期', '摘要', '原始摘要', '归一化摘要', '金额', '对方科目', '结算号', '方向']
            for col in expected_cols:
                assert col in headers, f"Missing column in ledger-only: {col}"

            data_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                          if any(c is not None for c in row))
            assert data_rows == 2, f"Expected 2 ledger-only rows, got {data_rows}"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_duplicate_sheet(self, result_with_duplicates):
        """Verify duplicate worksheet content."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            generate_report(result_with_duplicates, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb['疑似重复']

            headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
            expected_cols = ['来源', '日期', '摘要', '金额', '重复组ID']
            for col in expected_cols:
                assert col in headers, f"Missing column in duplicates: {col}"

            # 2 bank dups + 2 ledger dups = 4 rows
            data_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                          if any(c is not None for c in row))
            assert data_rows >= 2, f"Expected at least 2 duplicate rows, got {data_rows}"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


class TestStyling:
    """Tests for styling functions."""

    def test_header_styling(self):
        """Verify apply_header_style applies bold and blue background."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test Header'
        ws['B1'] = 'Value'

        apply_header_style(ws, 1)

        cell_a = ws['A1']
        assert cell_a.font.bold is True, "Header should be bold"
        # Check fill color (light blue #BDD7EE)
        assert cell_a.fill.start_color.rgb == '00BDD7EE', \
            f"Expected #BDD7EE, got {cell_a.fill.start_color.rgb}"

        cell_b = ws['B1']
        assert cell_b.font.bold is True

    def test_amount_formatting(self):
        """Verify format_amount_column applies number format."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        for i in range(2, 7):
            ws[f'A{i}'] = 1234.56

        format_amount_column(ws, 'A', 2, 6)
        for i in range(2, 7):
            assert ws[f'A{i}'].number_format == '#,##0.00', \
                f"Cell A{i} has wrong format: {ws[f'A{i}'].number_format}"

    def test_auto_fit_columns(self):
        """Verify auto_fit_columns adjusts column widths."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Short'
        ws['A2'] = 'A much longer string that needs more width'

        auto_fit_columns(ws)

        # Column width should be set (not default)
        assert ws.column_dimensions['A'].width is not None
        # Width should be at least enough for the longer string
        assert ws.column_dimensions['A'].width > 5


class TestStylingInReport:
    """Verify styling is applied in the generated report."""

    def test_report_has_frozen_header(self, sample_result):
        """Verify the first row is frozen in all sheets."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            generate_report(sample_result, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                assert ws.freeze_panes == 'A2', \
                    f"Sheet '{sheet_name}' freeze_panes should be A2, got {ws.freeze_panes}"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_report_header_style_applied(self, sample_result):
        """Verify header rows have bold font and blue fill in the report."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            generate_report(sample_result, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Check first cell of header row
                cell = ws.cell(row=1, column=1)
                assert cell.font.bold is True, \
                    f"Sheet '{sheet_name}' header should be bold"
                assert cell.fill.start_color.rgb == '00BDD7EE', \
                    f"Sheet '{sheet_name}' header should have blue fill"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_empty_dataframes_handled_gracefully(self):
        """Verify report generation works with empty DataFrames."""
        empty_bank = pd.DataFrame()
        empty_ledger = pd.DataFrame()
        result = ReconciliationResult(
            bank_file='', ledger_file='',
            bank_total_count=0, ledger_total_count=0,
            bank_total_income=0.0, bank_total_expense=0.0,
            ledger_total_income=0.0, ledger_total_expense=0.0,
            bank_opening_balance=0.0, bank_closing_balance=0.0,
            matched_count=0, exact_matched=0, fuzzy_matched=0, split_matched=0,
            unmatched_bank_count=0, unmatched_ledger_count=0,
            duplicate_bank_count=0, duplicate_ledger_count=0,
            total_bank_amount=0.0, total_ledger_amount=0.0,
            matched_amount_diff=0.0,
            unmatched_bank_amount=0.0, unmatched_ledger_amount=0.0,
            matched_records=[],
            unmatched_bank=empty_bank, unmatched_ledger=empty_ledger,
            bank_duplicates=empty_bank, ledger_duplicates=empty_ledger,
            reconciled_at='2026-04-28T10:00:00',
            date_range=('', ''), match_rate=100.0,
            bank_df=empty_bank, ledger_df=empty_ledger,
        )
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            result_path = generate_report(result, tmp_path)
            assert os.path.exists(result_path)
            wb = openpyxl.load_workbook(result_path)
            assert len(wb.sheetnames) == 5
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

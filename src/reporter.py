"""银行对账工具 — Excel 报告生成器

从 ReconciliationResult 生成包含 5 个工作表的格式化 .xlsx 对账报告。
"""

import os
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from pipeline import ReconciliationResult


# ═══════════════════════════════════════════════════════════════════════
# Style constants
# ═══════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
HEADER_FONT = Font(bold=True)
AMOUNT_FORMAT = '#,##0.00'
DATE_FORMAT = 'yyyy-mm-dd'


# ═══════════════════════════════════════════════════════════════════════
# Helper: ensure a value is safe for openpyxl (NaT → None, etc.)
# ═══════════════════════════════════════════════════════════════════════

def _safe_value(val):
    """Convert a value to something openpyxl can write.

    pd.NaT and pd.NA are converted to None.
    """
    if pd.isna(val):
        return None
    return val


# ═══════════════════════════════════════════════════════════════════════
# Styling functions
# ═══════════════════════════════════════════════════════════════════════

def apply_header_style(ws, row):
    """Apply header row style: bold font, light blue background.

    Args:
        ws: openpyxl Worksheet
        row: row number (1-indexed)
    """
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def format_amount_column(ws, col_letter, start_row, end_row):
    """Apply number format to an amount column.

    Args:
        ws: openpyxl Worksheet
        col_letter: column letter (e.g. 'E')
        start_row: first data row (1-indexed)
        end_row: last data row (1-indexed)
    """
    for row in range(start_row, end_row + 1):
        cell = ws[f'{col_letter}{row}']
        cell.number_format = AMOUNT_FORMAT


def auto_fit_columns(ws):
    """Auto-adjust column widths based on content.

    Args:
        ws: openpyxl Worksheet
    """
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            val = cell.value
            if val is not None:
                # Estimate width: CJK characters count ~2, others ~1
                str_val = str(val)
                length = 0
                for ch in str_val:
                    if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                        length += 2
                    else:
                        length += 1
                if length > max_length:
                    max_length = length

        # Set width with some padding, capped at 40
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 8)


# ═══════════════════════════════════════════════════════════════════════
# Sheet generators
# ═══════════════════════════════════════════════════════════════════════

def _write_summary_sheet(ws, result):
    """Write the 对账汇总 (Reconciliation Summary) sheet.

    Layout: two-column table with merged header rows for sections.
    """
    r = result  # shorthand

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30

    row = 1

    def write_row(label, value, is_header=False):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if is_header:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True, size=12)
        row += 1

    # ── Title ──
    write_row('银行对账报告', '', is_header=True)
    write_row('对账时间', r.reconciled_at[:19] if r.reconciled_at else '')
    write_row('银行文件', os.path.basename(r.bank_file) if r.bank_file else '')
    write_row('日记账文件', os.path.basename(r.ledger_file) if r.ledger_file else '')
    if r.date_range:
        write_row('日期范围', f"{r.date_range[0]} ~ {r.date_range[1]}")
    row += 1  # blank row

    # ── Bank summary ──
    write_row('银行统计', '', is_header=True)
    write_row('  总笔数', r.bank_total_count)
    write_row('  总收入', r.bank_total_income)
    write_row('  总支出', r.bank_total_expense)
    write_row('  净额', round(r.bank_total_income - r.bank_total_expense, 2))
    write_row('  期初余额', r.bank_opening_balance)
    write_row('  期末余额', r.bank_closing_balance)
    row += 1

    # ── Ledger summary ──
    write_row('日记账统计', '', is_header=True)
    write_row('  总笔数', r.ledger_total_count)
    write_row('  总收入', r.ledger_total_income)
    write_row('  总支出', r.ledger_total_expense)
    write_row('  净额', round(r.ledger_total_income - r.ledger_total_expense, 2))
    row += 1

    # ── Match statistics ──
    write_row('匹配统计', '', is_header=True)
    write_row('  精确匹配', r.exact_matched)
    write_row('  模糊匹配', r.fuzzy_matched)
    write_row('  拆分匹配', r.split_matched)
    write_row('  总匹配数', r.matched_count)
    write_row('  匹配率', f"{r.match_rate:.2f}%")
    row += 1

    # ── Differences ──
    write_row('差异分析', '', is_header=True)
    write_row('  银行独有笔数', r.unmatched_bank_count)
    write_row('  银行独有金额', r.unmatched_bank_amount)
    write_row('  日记账独有笔数', r.unmatched_ledger_count)
    write_row('  日记账独有金额', r.unmatched_ledger_amount)
    write_row('  已匹配金额差', r.matched_amount_diff)
    row += 1

    # ── Duplicates ──
    write_row('疑似重复', '', is_header=True)
    write_row('  银行重复笔数', r.duplicate_bank_count)
    write_row('  日记账重复笔数', r.duplicate_ledger_count)
    row += 1

    # ── Net difference ──
    write_row('金额差额', '', is_header=True)
    net_diff = round(r.total_bank_amount - r.total_ledger_amount, 2)
    write_row('  银行净额', r.total_bank_amount)
    write_row('  日记账净额', r.total_ledger_amount)
    write_row('  差额', net_diff)

    # Apply amount format to numeric cells in column B
    for r_idx in range(1, row):
        cell = ws.cell(row=r_idx, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = AMOUNT_FORMAT
            cell.alignment = Alignment(horizontal='right')

    # Apply header style to rows that are section headers
    for r_idx in range(1, row):
        cell_a = ws.cell(row=r_idx, column=1)
        if cell_a.font and cell_a.font.bold:
            # Already a header row — apply fill
            cell_a.fill = HEADER_FILL
            ws.cell(row=r_idx, column=2).fill = HEADER_FILL

    # Freeze the first row
    ws.freeze_panes = 'A2'


def _write_matched_detail_sheet(ws, result):
    """Write the 匹配明细 (Matched Detail) sheet.

    Lists all matched transactions with key fields from both sides.
    """
    headers = ['序号', '匹配类型', '银行日期', '银行摘要', '银行金额',
               '日记账日期', '日记账摘要', '日记账金额', '金额差', '日期差', '相似度分数']

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    apply_header_style(ws, 1)

    bank_df = result.bank_df
    ledger_df = result.ledger_df

    records = result.matched_records
    # Sort by match_type order: exact → fuzzy → split
    type_order = {'exact': 0, 'fuzzy': 1, 'split': 2}
    records = sorted(records, key=lambda r: type_order.get(r.match_type, 99))

    row = 2
    for seq, rec in enumerate(records, 1):
        # Bank-side data
        if bank_df is not None and len(bank_df) > 0 and rec.bank_idx < len(bank_df):
            bank_row = bank_df.iloc[rec.bank_idx]
            bank_date = _safe_value(bank_row.get('date', ''))
            bank_summary = bank_row.get('summary', '')
        else:
            bank_date = None
            bank_summary = ''

        # Ledger-side data
        if isinstance(rec.ledger_idx, list):
            # Split match: show the first ledger's info
            if ledger_df is not None and len(ledger_df) > 0:
                li0 = rec.ledger_idx[0] if rec.ledger_idx else 0
                if li0 < len(ledger_df):
                    ledger_row = ledger_df.iloc[li0]
                    ledger_date = _safe_value(ledger_row.get('date', ''))
                    ledger_summary = f"{ledger_row.get('summary', '')} (...等{len(rec.ledger_idx)}笔)"
                else:
                    ledger_date = None
                    ledger_summary = ''
            else:
                ledger_date = None
                ledger_summary = ''
        else:
            if ledger_df is not None and len(ledger_df) > 0 and rec.ledger_idx < len(ledger_df):
                ledger_row = ledger_df.iloc[rec.ledger_idx]
                ledger_date = _safe_value(ledger_row.get('date', ''))
                ledger_summary = ledger_row.get('summary', '')
            else:
                ledger_date = None
                ledger_summary = ''

        # Format dates
        if hasattr(bank_date, 'strftime'):
            bank_date_str = bank_date.strftime('%Y-%m-%d')
        else:
            bank_date_str = str(bank_date) if bank_date else ''

        if hasattr(ledger_date, 'strftime'):
            ledger_date_str = ledger_date.strftime('%Y-%m-%d')
        else:
            ledger_date_str = str(ledger_date) if ledger_date else ''

        ws.cell(row=row, column=1, value=seq)
        ws.cell(row=row, column=2, value=rec.match_type)
        ws.cell(row=row, column=3, value=bank_date_str)
        ws.cell(row=row, column=4, value=str(bank_summary))
        ws.cell(row=row, column=5, value=rec.bank_amount)
        ws.cell(row=row, column=6, value=ledger_date_str)
        ws.cell(row=row, column=7, value=str(ledger_summary))
        ws.cell(row=row, column=8, value=rec.ledger_amount)
        ws.cell(row=row, column=9, value=rec.amount_diff)
        ws.cell(row=row, column=10, value=rec.date_diff)
        ws.cell(row=row, column=11, value=rec.score)

        row += 1

    # Format amount columns (E, H, I)
    if row > 2:
        format_amount_column(ws, 'E', 2, row - 1)
        format_amount_column(ws, 'H', 2, row - 1)
        format_amount_column(ws, 'I', 2, row - 1)

    # Auto-fit columns
    auto_fit_columns(ws)

    # Freeze header
    ws.freeze_panes = 'A2'


def _write_bank_only_sheet(ws, result):
    """Write the 银行独有 (Bank Only) sheet.

    Transactions that appear only in the bank statement.
    """
    headers = ['日期', '摘要', '原始摘要', '归一化摘要', '金额', '余额',
               '对方名称', '对方账号']

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    apply_header_style(ws, 1)

    df = result.unmatched_bank
    if df is None or len(df) == 0:
        auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
        return

    row = 2
    for _, txn in df.iterrows():
        date_val = _safe_value(txn.get('date', ''))
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val) if date_val else ''

        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=str(txn.get('summary', '')))
        ws.cell(row=row, column=3, value=str(txn.get('summary', '')))
        ws.cell(row=row, column=4, value=str(txn.get('normalized_summary', '')))
        ws.cell(row=row, column=5, value=txn.get('normalized_amount', 0))
        ws.cell(row=row, column=6, value=txn.get('balance', 0))
        ws.cell(row=row, column=7, value=str(txn.get('counterparty', '')))
        ws.cell(row=row, column=8, value=str(txn.get('counterparty_acct', '')))

        row += 1

    # Format amount column (E)
    if row > 2:
        format_amount_column(ws, 'E', 2, row - 1)

    auto_fit_columns(ws)
    ws.freeze_panes = 'A2'


def _write_ledger_only_sheet(ws, result):
    """Write the 日记账独有 (Ledger Only) sheet.

    Transactions that appear only in the company ledger.
    """
    headers = ['日期', '摘要', '原始摘要', '归一化摘要', '金额',
               '对方科目', '结算号', '方向']

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    apply_header_style(ws, 1)

    df = result.unmatched_ledger
    if df is None or len(df) == 0:
        auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
        return

    row = 2
    for _, txn in df.iterrows():
        date_val = _safe_value(txn.get('date', ''))
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val) if date_val else ''

        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=str(txn.get('summary', '')))
        ws.cell(row=row, column=3, value=str(txn.get('summary', '')))
        ws.cell(row=row, column=4, value=str(txn.get('normalized_summary', '')))
        ws.cell(row=row, column=5, value=txn.get('normalized_amount', 0))
        ws.cell(row=row, column=6, value=str(txn.get('counterparty_subject', '')))
        ws.cell(row=row, column=7, value=str(txn.get('voucher_no', '')))
        ws.cell(row=row, column=8, value=str(txn.get('direction', '')))

        row += 1

    # Format amount column (E)
    if row > 2:
        format_amount_column(ws, 'E', 2, row - 1)

    auto_fit_columns(ws)
    ws.freeze_panes = 'A2'


def _write_duplicate_sheet(ws, result):
    """Write the 疑似重复 (Suspected Duplicates) sheet.

    Lists duplicate transactions from both bank and ledger sides.
    """
    headers = ['来源', '日期', '摘要', '金额', '重复组ID']

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    apply_header_style(ws, 1)

    row = 2

    # Bank duplicates
    bank_dups = result.bank_duplicates
    if bank_dups is not None and len(bank_dups) > 0:
        for _, txn in bank_dups.iterrows():
            date_val = _safe_value(txn.get('date', ''))
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val) if date_val else ''

            ws.cell(row=row, column=1, value='银行')
            ws.cell(row=row, column=2, value=date_str)
            ws.cell(row=row, column=3, value=str(txn.get('summary', '')))
            ws.cell(row=row, column=4, value=txn.get('normalized_amount', 0))
            ws.cell(row=row, column=5, value=int(txn.get('duplicate_group_id', -1)))
            row += 1

    # Ledger duplicates
    ledger_dups = result.ledger_duplicates
    if ledger_dups is not None and len(ledger_dups) > 0:
        for _, txn in ledger_dups.iterrows():
            date_val = _safe_value(txn.get('date', ''))
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val) if date_val else ''

            ws.cell(row=row, column=1, value='日记账')
            ws.cell(row=row, column=2, value=date_str)
            ws.cell(row=row, column=3, value=str(txn.get('summary', '')))
            ws.cell(row=row, column=4, value=txn.get('normalized_amount', 0))
            ws.cell(row=row, column=5, value=int(txn.get('duplicate_group_id', -1)))
            row += 1

    # Format amount column (D)
    if row > 2:
        format_amount_column(ws, 'D', 2, row - 1)

    auto_fit_columns(ws)
    ws.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════════════════════════
# Main entry point
# ═══════════════════════════════════════════════════════════════════════

def generate_report(result, output_path):
    """Generate a formatted .xlsx reconciliation report with 5 worksheets.

    Sheets:
        1. 对账汇总 — Summary statistics
        2. 匹配明细 — Matched transaction details
        3. 银行独有 — Bank-only transactions
        4. 日记账独有 — Ledger-only transactions
        5. 疑似重复 — Suspected duplicates

    Args:
        result: ReconciliationResult from pipeline.run_reconciliation()
        output_path: Path for the output .xlsx file

    Returns:
        str: The output file path
    """
    if not isinstance(result, ReconciliationResult):
        raise TypeError(
            f"Expected ReconciliationResult, got {type(result).__name__}"
        )

    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: 对账汇总
    ws1 = wb.create_sheet('对账汇总')
    _write_summary_sheet(ws1, result)

    # Sheet 2: 匹配明细
    ws2 = wb.create_sheet('匹配明细')
    _write_matched_detail_sheet(ws2, result)

    # Sheet 3: 银行独有
    ws3 = wb.create_sheet('银行独有')
    _write_bank_only_sheet(ws3, result)

    # Sheet 4: 日记账独有
    ws4 = wb.create_sheet('日记账独有')
    _write_ledger_only_sheet(ws4, result)

    # Sheet 5: 疑似重复
    ws5 = wb.create_sheet('疑似重复')
    _write_duplicate_sheet(ws5, result)

    # Save
    wb.save(output_path)
    return output_path

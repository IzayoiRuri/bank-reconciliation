"""银行对账工具 — 文件解析器

解析银行流水(.xlsx)和公司日记账(.xls)为统一格式的 DataFrame。
支持多种银行格式：招商银行（标准 xlsx）、工商银行（对账单）。
"""

import pandas as pd
import openpyxl
import xlrd
import re
from datetime import datetime

from config import BANK_COLUMN_ALIASES, LEDGER_COLUMN_ALIASES, BANK_FORMATS


# ═══════════════════════════════════════════════════════════════════════
# Column auto-detection
# ═══════════════════════════════════════════════════════════════════════

def auto_detect_columns(df, aliases_dict):
    """根据列名别名映射自动匹配 DataFrame 的实际列名到标准列名。

    Args:
        df: pandas DataFrame
        aliases_dict: {标准列名: [别名列表]}

    Returns:
        dict: {标准列名: 实际列名}，未匹配到的值为 None
    """
    result = {}
    actual_cols = [str(c).strip() for c in df.columns]

    for std_name, aliases in aliases_dict.items():
        matched = None
        for alias in aliases:
            if alias in actual_cols:
                matched = alias
                break
        result[std_name] = matched

    return result


# ═══════════════════════════════════════════════════════════════════════
# Bank statement parser (.xlsx) — dispatch
# ═══════════════════════════════════════════════════════════════════════

def parse_bank_statement(filepath, bank_format='auto'):
    """解析银行流水文件 (.xlsx)。

    自动识别列名并标准化输出：
    date, debit, credit, balance, summary, counterparty, counterparty_acct

    Args:
        filepath: .xlsx 文件路径
        bank_format: 'auto' (自动检测), 'zhaoshang' (招商银行), 'gonghang' (工商银行对账单)

    Returns:
        pandas DataFrame
    """
    if bank_format == 'gonghang':
        return _parse_bank_gonghang(filepath)
    elif bank_format == 'jianshe':
        return _parse_bank_historydetail(filepath)
    elif bank_format == 'zhaoshang':
        return _parse_bank_standard(filepath, header_row=1)
    else:
        # auto: try standard → gonghang → historydetail
        result = _parse_bank_standard(filepath)
        if len(result) > 0:
            return result
        result = _parse_bank_gonghang(filepath)
        if len(result) > 0:
            return result
        return _parse_bank_historydetail(filepath)


# ═══════════════════════════════════════════════════════════════════════
# Bank parser: 招商银行 (standard xlsx, row 1 = header)
# ═══════════════════════════════════════════════════════════════════════

def _find_bank_header_row(ws):
    """Find the header row in a bank statement worksheet.

    Scans rows looking for one that contains common bank statement column names
    like '交易日', '借方金额', etc. Returns the 1-indexed row number.
    """
    header_keywords = ['交易日', '交易时间', '借方金额', '贷方金额']

    for row_idx in range(1, ws.max_row + 1):
        row_vals = [str(cell.value) if cell.value is not None else '' for cell in ws[row_idx]]
        row_text = ' '.join(row_vals)

        matches = sum(1 for kw in header_keywords if kw in row_text)
        if matches >= 3:
            return row_idx

    return 1


def _build_standardized_df(df, col_map):
    """Build standardized DataFrame from a raw parsed DataFrame and column map.

    Args:
        df: raw DataFrame with original column names
        col_map: {标准列名: 实际列名} from auto_detect_columns

    Returns:
        standardized DataFrame (date, debit, credit, balance, summary, counterparty, counterparty_acct)
    """
    result = pd.DataFrame()

    # Date
    date_col = col_map.get('date')
    if date_col and date_col in df.columns:
        result['date'] = pd.to_datetime(df[date_col], errors='coerce')
    else:
        result['date'] = pd.NaT

    # Debit
    debit_col = col_map.get('debit')
    if debit_col and debit_col in df.columns:
        result['debit'] = pd.to_numeric(df[debit_col], errors='coerce').fillna(0)
    else:
        result['debit'] = 0.0

    # Credit
    credit_col = col_map.get('credit')
    if credit_col and credit_col in df.columns:
        result['credit'] = pd.to_numeric(df[credit_col], errors='coerce').fillna(0)
    else:
        result['credit'] = 0.0

    # Balance
    balance_col = col_map.get('balance')
    if balance_col and balance_col in df.columns:
        result['balance'] = pd.to_numeric(df[balance_col], errors='coerce').fillna(0)
    else:
        result['balance'] = 0.0

    # Summary
    summary_col = col_map.get('summary')
    if summary_col and summary_col in df.columns:
        result['summary'] = df[summary_col].fillna('').astype(str)
    else:
        result['summary'] = ''

    # Counterparty
    cp_col = col_map.get('counterparty')
    if cp_col and cp_col in df.columns:
        result['counterparty'] = df[cp_col].fillna('').astype(str)
    else:
        result['counterparty'] = ''

    # Counterparty account
    cp_acct_col = col_map.get('counterparty_acct')
    if cp_acct_col and cp_acct_col in df.columns:
        result['counterparty_acct'] = df[cp_acct_col].fillna('').astype(str)
    else:
        result['counterparty_acct'] = ''

    # Drop rows where date is NaT (metadata/empty rows)
    result = result.dropna(subset=['date']).reset_index(drop=True)

    # Fill any remaining NaN amounts with 0
    for col in ['debit', 'credit', 'balance']:
        result[col] = result[col].fillna(0)

    return result


def _parse_bank_standard(filepath, header_row=None):
    """Parse standard bank xlsx via openpyxl (招商银行 format).

    Args:
        filepath: .xlsx file path
        header_row: force header row (1-indexed), None=auto-detect

    Returns:
        standardized DataFrame
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    if header_row is None:
        header_row = _find_bank_header_row(ws)

    # Read header values
    header_vals = []
    for cell in ws[header_row]:
        val = cell.value
        header_vals.append(str(val).strip() if val is not None else '')

    # Read all data rows after header
    data_rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_vals = []
        for cell in ws[row_idx]:
            row_vals.append(cell.value)
        data_rows.append(row_vals)

    df = pd.DataFrame(data_rows, columns=header_vals)
    col_map = auto_detect_columns(df, BANK_COLUMN_ALIASES)
    return _build_standardized_df(df, col_map)


# ═══════════════════════════════════════════════════════════════════════
# Bank parser: 工商银行 (对账单 format, 36-column wide table)
# ═══════════════════════════════════════════════════════════════════════

# Gonghang (ICBC) 对账单 column index → standard field mapping
_GONGHANG_COLUMN_MAP = {
    3: 'date',               # 交易日
    7: 'debit',              # 借方金额
    8: 'credit',             # 贷方金额
    9: 'balance',            # 余额
    10: 'summary',           # 摘要
    19: 'counterparty',      # 收(付)方名称
    20: 'counterparty_acct', # 收(付)方账号
}

# Row index where the column header row lives (0-indexed)
_GONGHANG_HEADER_ROW = 12

# Row index where data starts (0-indexed, after header)
_GONGHANG_DATA_START = 13


def _parse_bank_gonghang(filepath):
    """Parse ICBC 对账单 format bank statement.

    The file has:
      - Rows 0-11: metadata block (标题/接口版本/账户信息 etc.)
      - Row 12: column headers (36 columns)
      - Rows 13+: transaction data

    Uses pandas read_excel with header=None because openpyxl's dimension
    detection does not see all 36 columns of this format.

    Args:
        filepath: .xlsx file path

    Returns:
        standardized DataFrame (date, debit, credit, balance, summary, counterparty, counterparty_acct)
    """
    # Read all rows with no header — pandas sees all 36 columns correctly
    raw = pd.read_excel(filepath, header=None, engine='openpyxl')

    # Extract column headers from the designated header row
    header_vals = []
    for c in range(raw.shape[1]):
        val = raw.iloc[_GONGHANG_HEADER_ROW, c]
        header_vals.append(str(val).strip() if pd.notna(val) else f'_col{c}')

    # Extract data rows (after header)
    data = raw.iloc[_GONGHANG_DATA_START:].copy()
    data.columns = header_vals
    data = data.reset_index(drop=True)

    # Build standardized result directly via fixed column index mapping
    result = pd.DataFrame()

    for col_idx, std_field in _GONGHANG_COLUMN_MAP.items():
        if col_idx < raw.shape[1]:
            series = raw.iloc[_GONGHANG_DATA_START:, col_idx].reset_index(drop=True)

            if std_field == 'date':
                result['date'] = pd.to_datetime(series, errors='coerce')
            elif std_field in ('debit', 'credit', 'balance'):
                result[std_field] = pd.to_numeric(series, errors='coerce').fillna(0)
            else:
                result[std_field] = series.fillna('').astype(str)
        else:
            if std_field == 'date':
                result['date'] = pd.NaT
            elif std_field in ('debit', 'credit', 'balance'):
                result[std_field] = 0.0
            else:
                result[std_field] = ''

    # Drop rows where date is NaT (should be none, but safety)
    result = result.dropna(subset=['date']).reset_index(drop=True)

    for col in ['debit', 'credit', 'balance']:
        result[col] = result[col].fillna(0)

    return result


# ═══════════════════════════════════════════════════════════════════════
# Bank parser: 建设银行 (HISTORYDETAIL format, 借贷标志 + 发生额)
# ═══════════════════════════════════════════════════════════════════════

def _parse_bank_historydetail(filepath):
    """Parse HISTORYDETAIL format bank statement (工商银行/建设银行 etc).

    Format:
      - Row 0: [HISTORYDETAIL] tag
      - Row 1: column headers (凭证号/对方账号/交易时间/借贷标志/对方单位/...)
      - Rows 2+: transaction data
      - Amount: single `发生额` column + `借贷标志` (借/贷) to determine direction
        - 借 (debit) → 支出, 贷 (credit) → 收入

    Uses pandas read_excel with header=None for reliable column access.

    Args:
        filepath: .xlsx file path

    Returns:
        standardized DataFrame (date, debit, credit, balance, summary, counterparty, counterparty_acct)
    """
    raw = pd.read_excel(filepath, header=None, engine='openpyxl')

    # Column index mapping (0-indexed)
    #  0: 凭证号      1: 对方账号      2: 交易时间
    #  3: 借贷标志    4: 对方单位      5: 对方行号
    #  6: 用途        7: 摘要          8: 附言
    #  9: 回单个性化信息  10: 发生额    11: 入账日期
    # 12: 余额       13: 入账时间      14: 本方账号
    # 15: 转出金额   16: 转入金额      17: 币种

    # Data starts at row 2 (0-indexed), skip [HISTORYDETAIL] + header
    data = raw.iloc[2:].copy()
    data = data.reset_index(drop=True)

    # Parse amount from 发生额 (col 10) — may have commas
    amount_col = data.iloc[:, 10] if raw.shape[1] > 10 else pd.Series([pd.NA] * len(data))
    amount_str = amount_col.fillna('0').astype(str).str.replace(',', '').str.strip()
    amounts = pd.to_numeric(amount_str, errors='coerce').fillna(0)

    # Direction from 借贷标志 (col 3)
    direction_col = data.iloc[:, 3] if raw.shape[1] > 3 else pd.Series([''] * len(data))
    direction = direction_col.fillna('').astype(str).str.strip()

    # Build standardized result
    result = pd.DataFrame()

    # Date: prefer 入账日期 (col 11, "YYYY-MM-DD"), fall back to 交易时间 (col 2)
    date_col = data.iloc[:, 11] if raw.shape[1] > 11 else pd.Series([pd.NaT] * len(data))
    date_fallback = data.iloc[:, 2] if raw.shape[1] > 2 else pd.Series([pd.NaT] * len(data))
    date_vals = date_col.fillna(date_fallback).astype(str).str[:10]
    result['date'] = pd.to_datetime(date_vals, errors='coerce')

    # Debit: 发生额 where 借贷标志='借', else 0
    is_debit = direction == '借'
    result['debit'] = amounts.where(is_debit, 0)

    # Credit: 发生额 where 借贷标志='贷', else 0
    is_credit = direction == '贷'
    result['credit'] = amounts.where(is_credit, 0)

    # Balance (col 12)
    balance_col = data.iloc[:, 12] if raw.shape[1] > 12 else pd.Series([0] * len(data))
    balance_str = balance_col.fillna('0').astype(str).str.replace(',', '').str.strip()
    result['balance'] = pd.to_numeric(balance_str, errors='coerce').fillna(0)

    # Summary: 用途 (col 6), fall back to 摘要 (col 7)
    usage_col = data.iloc[:, 6] if raw.shape[1] > 6 else pd.Series([''] * len(data))
    summary_col = data.iloc[:, 7] if raw.shape[1] > 7 else pd.Series([''] * len(data))
    result['summary'] = usage_col.fillna('').astype(str)
    # If 用途 is empty, use 摘要
    empty_mask = result['summary'].str.strip() == ''
    result.loc[empty_mask, 'summary'] = summary_col.fillna('').astype(str)[empty_mask]

    # Counterparty: 对方单位 (col 4)
    cp_col = data.iloc[:, 4] if raw.shape[1] > 4 else pd.Series([''] * len(data))
    result['counterparty'] = cp_col.fillna('').astype(str)

    # Counterparty account: 对方账号 (col 1)
    cp_acct_col = data.iloc[:, 1] if raw.shape[1] > 1 else pd.Series([''] * len(data))
    result['counterparty_acct'] = cp_acct_col.fillna('').astype(str)

    # Cleanup
    result = result.dropna(subset=['date']).reset_index(drop=True)
    for col in ['debit', 'credit', 'balance']:
        result[col] = result[col].fillna(0)

    return result


# ═══════════════════════════════════════════════════════════════════════
# Ledger parser (.xls)
# ═══════════════════════════════════════════════════════════════════════

def _is_summary_row(row_values):
    """Check if a ledger row is a summary row (本日合计/本月合计/本年累计/当前合计/当前累计/上年结转).

    Args:
        row_values: list of cell values for a row

    Returns:
        bool: True if row should be filtered out
    """
    row_str = '|'.join([str(v) if v is not None else '' for v in row_values])

    summary_keywords = ['合计', '累计', '上年结转']
    for kw in summary_keywords:
        if kw in row_str:
            return True

    return False


def parse_ledger(filepath):
    """解析公司日记账文件 (.xls)。

    过滤日合计/月合计/累计行，合并年/月/日为 date 列。
    标准化输出：
    date, summary, voucher_no, counterparty_subject, debit, credit, direction, balance

    Args:
        filepath: .xls 文件路径

    Returns:
        pandas DataFrame
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]

    data_rows = []
    for r in range(1, ws.nrows):
        row_vals = [ws.cell_value(r, c) for c in range(ws.ncols)]

        if _is_summary_row(row_vals):
            continue

        data_rows.append(row_vals)

    df = pd.DataFrame(data_rows, columns=headers)
    col_map = auto_detect_columns(df, LEDGER_COLUMN_ALIASES)

    result = pd.DataFrame()

    # Combine year, month, day into date
    year_col = col_map.get('year')
    month_col = col_map.get('month')
    day_col = col_map.get('day')

    year_vals = pd.to_numeric(df[year_col], errors='coerce').fillna(1).astype(int) if year_col else 2026
    month_vals = pd.to_numeric(df[month_col], errors='coerce').fillna(1).astype(int) if month_col else 1
    day_vals = pd.to_numeric(df[day_col], errors='coerce').fillna(1).astype(int) if day_col else 1

    def build_date(y, m, d):
        try:
            return pd.Timestamp(year=int(y), month=int(m), day=int(d))
        except (ValueError, TypeError):
            return pd.NaT

    dates = []
    for i in range(len(df)):
        dates.append(build_date(
            year_vals.iloc[i] if hasattr(year_vals, 'iloc') else year_vals[i] if isinstance(year_vals, (list, pd.Series)) else year_vals,
            month_vals.iloc[i] if hasattr(month_vals, 'iloc') else month_vals[i] if isinstance(month_vals, (list, pd.Series)) else month_vals,
            day_vals.iloc[i] if hasattr(day_vals, 'iloc') else day_vals[i] if isinstance(day_vals, (list, pd.Series)) else day_vals))

    result['date'] = dates

    # Summary
    summary_col = col_map.get('summary')
    if summary_col and summary_col in df.columns:
        result['summary'] = df[summary_col].fillna('').astype(str)
    else:
        result['summary'] = ''

    # Voucher number
    voucher_col = col_map.get('voucher_no')
    if voucher_col and voucher_col in df.columns:
        result['voucher_no'] = df[voucher_col].fillna('').astype(str)
    else:
        result['voucher_no'] = ''

    # Counterparty subject
    cp_subj_col = col_map.get('counterparty_subject')
    if cp_subj_col and cp_subj_col in df.columns:
        result['counterparty_subject'] = df[cp_subj_col].fillna('').astype(str)
    else:
        result['counterparty_subject'] = ''

    # Debit
    debit_col = col_map.get('debit')
    if debit_col and debit_col in df.columns:
        result['debit'] = pd.to_numeric(df[debit_col], errors='coerce').fillna(0)
    else:
        result['debit'] = 0.0

    # Credit
    credit_col = col_map.get('credit')
    if credit_col and credit_col in df.columns:
        result['credit'] = pd.to_numeric(df[credit_col], errors='coerce').fillna(0)
    else:
        result['credit'] = 0.0

    # Direction
    direction_col = col_map.get('direction')
    if direction_col and direction_col in df.columns:
        result['direction'] = df[direction_col].fillna('').astype(str)
    else:
        result['direction'] = ''

    # Balance
    balance_col = col_map.get('balance')
    if balance_col and balance_col in df.columns:
        result['balance'] = pd.to_numeric(df[balance_col], errors='coerce').fillna(0)
    else:
        result['balance'] = 0.0

    result = result.dropna(subset=['date']).reset_index(drop=True)

    for col in ['debit', 'credit', 'balance']:
        result[col] = result[col].fillna(0)

    return result


# ═══════════════════════════════════════════════════════════════════════
# Transaction normalization
# ═══════════════════════════════════════════════════════════════════════

def normalize_transaction(df, source):
    """统一标准化交易记录。

    - 添加 source 列
    - 添加 normalized_amount（统一符号：正数=收入，负数=支出）
      - 银行：贷方 - 借方
      - 日记账：借方 - 贷方
    - 添加 month 列（YYYY-MM）
    - 清理空值和异常值

    Args:
        df: parse_bank_statement 或 parse_ledger 输出的 DataFrame
        source: 'bank' 或 'ledger'

    Returns:
        标准化后的 DataFrame（添加了 source, normalized_amount, month 列）
    """
    df = df.copy()

    df['source'] = source

    if source == 'bank':
        df['normalized_amount'] = df['credit'] - df['debit']
    elif source == 'ledger':
        df['normalized_amount'] = df['debit'] - df['credit']
    else:
        raise ValueError(f"Unknown source: {source}. Must be 'bank' or 'ledger'.")

    if 'date' in df.columns:
        df['month'] = df['date'].apply(
            lambda d: d.strftime('%Y-%m') if pd.notna(d) else ''
        )
    else:
        df['month'] = ''

    return df


# ═══════════════════════════════════════════════════════════════════════
# Main entry point
# ═══════════════════════════════════════════════════════════════════════

def load_and_parse(bank_path, ledger_path, bank_format='auto'):
    """加载并解析银行流水和公司日记账。

    Args:
        bank_path: 银行流水 .xlsx 文件路径
        ledger_path: 公司日记账 .xls 文件路径
        bank_format: 'auto', 'zhaoshang', 或 'gonghang'

    Returns:
        (bank_df, ledger_df): 标准化后的两个 DataFrame
    """
    bank_df = parse_bank_statement(bank_path, bank_format=bank_format)
    ledger_df = parse_ledger(ledger_path)

    bank_df = normalize_transaction(bank_df, 'bank')
    ledger_df = normalize_transaction(ledger_df, 'ledger')

    return bank_df, ledger_df
